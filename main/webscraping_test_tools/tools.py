import re
import json
from website_tools.company_website import website_info
from general_tools.tools import json2composite
from openpyxl import Workbook, load_workbook

def load_samples(path):
    return json.loads(open(path, mode="r", encoding="utf-8").read())


def text_out_test(file_name, country, sample_path, language):
    counter = 0
    not_succeed_domains = 0
    full_empty_results = 0
    samples = load_samples(sample_path)

    outfile = open(file_name, mode="w", encoding="utf-8")
    outfile.write("[\n")
    outfile.close()
    for sample in samples[300: 305]:
        counter += 1
        empty = True
        outfile = open(file_name, mode="a", encoding="utf-8")
        outfile.write("    {\n")
        outfile.write('        "domain":"' + sample["Website"] + '"\n')
        result = website_info(sample["Website"], sample["Organization Name"], language, country=country)
        if(result):
            if(result["result"]["addresses"]):
                outfile.write('        "addresses":[\n')
                for add in result["result"]["addresses"]:
                    outfile.write('            ' + str(add) + ',\n')
                outfile.write('        ],\n')
            else:
                outfile.write('        "addresses":[],\n')
            
            if(result["result"]["phones"]):
                outfile.write('        "phones":[\n')
                for phone in result["result"]["phones"]:
                    outfile.write('            "' + phone + '",\n')
                outfile.write('        ],\n')
            else:
                outfile.write('        "phones":[],\n')
            
            if(result["result"]["emails"]):
                outfile.write('        "emails":[\n')
                for email in result["result"]["emails"]:
                    outfile.write('            "' + email + '",\n')
                outfile.write('        ],\n')
            else:
                outfile.write('        "emails":[],\n')

            if("logo_url" in result["result"].keys()):
                outfile.write('        "logo_url":' + result["result"]["logo_url"] + '\n')
            
            if(result["result"]["social_pages"]):
                outfile.write('        "social_pages":[\n')
                for k in result["result"]["social_pages"].keys():
                    outfile.write('            "' + k + '":"' + result["result"]["social_pages"][k] + '",\n')
                outfile.write('        ],\n')
        
            for k in result["result"].keys():
                if(result["result"][k]):
                    empty = False
                    break
            if(empty):
               full_empty_results += 1
        else:
            not_succeed_domains += 1
                
        outfile.write("    },\n")
        outfile.close()
        print("counter >>> ", counter, " - not_succeed_domains >>> ", not_succeed_domains, " - full empty result >>> ", full_empty_results)
        print(100 * "*")

    outfile = open(file_name, mode="a", encoding="utf-8")
    outfile.write("]")
    outfile.close()

def json_out_test(file_name, country, sample_path, language):
    samples = load_samples(sample_path)
    all_data = []
    for sample in samples[9:10]:
        all_data.append(website_info(sample["Website"], sample["Organization Name"], language, country=country))
        with open(file_name, mode="w", encoding="utf-8") as outfile:
            outfile.write(json.dumps(all_data, indent=4, ensure_ascii=False))


def test_domain(domain, org_name, country, language):
    result = website_info(domain, org_name, language, country)
    #print(result)
    #for add in result["result"]["addresses"]:
    #    print(add)
    with open("temp.json", mode="w", encoding="utf-8") as outfile:
        outfile.write(json.dumps(result, ensure_ascii=False, indent=4))

def test_regex(regex, text, find_all=False):
    if(find_all):
        items = re.findall(regex, text, flags=re.IGNORECASE)
        print(len(items), " >>> ")
        for item in items:
            print(item[0])

    else:
        m = re.search(regex, text, flags=re.IGNORECASE)
        if(m):
            print(m.group(0))
        else:
            print("not matched ...")

def test_addresses(regex, sample_path):
    counter = 0
    p_match = 0
    adds = load_samples(sample_path)
    for add in adds[:]:
        add = re.sub("\s{2,}", " ", add)
        add = re.sub("\(|\)", " ", add)
        add = add.strip()
        m = re.search(regex, add, flags=re.IGNORECASE)
        if(not m):
            print("Not matched >>> ", add)
            counter += 1
        elif(m.group(0) != add):
            print(add)
            print("P matched >>> ", m.group(0))
            print(50 * "*")
            p_match += 1
    print("all >>> ", len(adds))
    print("No Match >>> ", counter)
    print("P Match >>> ", p_match)

def domains2url(sample_path):
    f = open("urls.txt", mode="w", encoding="utf-8")
    f.write("[\n")
    samples = load_samples(sample_path)
    for sample in samples:
        f.write('    "http://' + sample["Website"] + '",\n\n')
    f.write("]")
    f.close()

def excel_out_test(dest_filename, country, sample_path, language):
    counter = 0
    not_succeed_domains = 0
    samples = load_samples(sample_path)

    #wb = Workbook()
    wb = load_workbook(dest_filename)
    first_ws = wb.active
    first_ws.title = "output data"

    fields = {
        "web_title":4,
        "addresses": 5,
        "google_address":6,
        "phones": 7,
        "emails":8
    }
    social_fields = {
        "facebook": 9,
        "instagram": 10,
        "linkedin": 11,
        "twitter": 12
    }

    first_ws.cell(row=1, column=1, value="domain")
    first_ws.cell(row=1, column=2, value="org_name")
    first_ws.cell(row=1, column=3, value="domain_valid")

    for k in fields.keys():
        first_ws.cell(row=1, column=fields[k], value=k)
    for k in social_fields.keys():
        first_ws.cell(row=1, column=social_fields[k], value=k)
    for index, sample in enumerate(samples[50:350]):
        counter += 1
        first_ws.cell(row=index + 51, column=1, value=sample["Website"])
        first_ws.cell(row=index + 51, column=2, value=sample["Organization Name"])

        result = website_info(sample["Website"], sample["Organization Name"], language, country=country)
        if(result):
            first_ws.cell(row=index + 51, column=3, value="valid")
            for key in fields.keys():
                if(result["result"].get(key)):
                    first_ws.cell(row=index + 51, column=fields[key], value=str(result["result"][key]))
                #else:
                #    first_ws.cell(row=index + 51, column=fields[key], value="-")
            if(result["result"].get("social_pages")):
                for key in social_fields.keys():
                    if(result["result"]["social_pages"].get(key)):
                        first_ws.cell(row=index + 51, column=social_fields[key], value=result["result"]["social_pages"][key])
                    #else:
                    #    first_ws.cell(row=index + 51, column=social_fields[key], value="-")

        else:
            not_succeed_domains += 1
        wb.save(dest_filename)     
        print("counter >>> ", counter, " - not_succeed_domains >>> ", not_succeed_domains)
        print(100 * "*")

def test_composite(out_put_path, sample_path, country):
    samples = load_samples(sample_path)
    all_data = []
    for obj in samples[:1]:
        print(obj["input_data"]["domain"])
        all_data.append(json2composite(obj, country))
    with open(out_put_path, mode="w", encoding="utf-8") as outfile:
        outfile.write(json.dumps(all_data, ensure_ascii=False, indent=4))